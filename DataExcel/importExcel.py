import numpy as np
import datetime
import openpyxl
from openpyxl import Workbook
import matplotlib.pyplot as plt
wb = openpyxl.load_workbook('./logTest.xlsx')

ws = wb.active
# найдем количество столбцов
max_col = ws.max_column
print("количество столбцов", max_col)
# количество ячеек
id = ws.max_row
print("количество ячеек", id)

# Добавление и перезапись ячеек, не забыть что надо будет сдвиг сделать на 1
# создадим словарь заголовков
nameColumn = [{}]
for i in range(1, max_col):
    nameColumn.append(dict([('name', ws.cell(row=1, column=i).value)]))
    nameColumn = list(filter(None, nameColumn))

nameParametr = 'Y3 (023)'
for i in range(0, len(nameColumn)):
    if nameColumn[i]['name'] == nameParametr:
        # сдвигаем на 1 вперед, чтобы потом не проворонить все с столбцами
        A = i + 1
        break
    i = i + 1
print(A, '   ', nameColumn[A]['name'])

# сохраним переменные в массив
# счетчик времени
timeCounter = np.zeros(id - 1)
# рабочая константа
workParametr = np.zeros(id - 1)

for i in range(0, id - 1):
    timeCounter[i] = i * 200

    workParametr[i] = ws.cell(row=i + 2, column=A).value
wb.close()

print(workParametr)

# сгенерируем новый excel file

wb=Workbook()
ws=wb.active
ws.title="formul sheet"

ws.cell(row=1,column=1,value='Time')
ws.cell(row=1,column=2,value=nameColumn[A]['name'])
for i in range(0,len(workParametr)):
    ws.cell(row=i+2,column=1,value=timeCounter[i])
    ws.cell(row=i+2,column=2,value=workParametr[i])


wb.save(filename='test.xlsx')

plt.plot(timeCounter,workParametr)
plt.grid()

now = datetime.datetime.now()
now = str(now)
for i in range(len(now)):
    if now[i] == ':':
        now = now.replace(now[i], '-')
    if now[i] == ' ':
        now = now.replace(now[i], '_')

now = now + '.png'

plt.grid()
plt.savefig(now, format='png')